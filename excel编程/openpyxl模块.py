#!/usr/bin/env python
# -*- coding: utf-8 -*-
#from __future__ import print_function
import openpyxl

wb = openpyxl.load_workbook(r'D:\pythonproject\coreapp3\excel编程\5月31日体育节目表.xlsx'.decode('utf-8'))
#加载一个存在的excel文档使用load_workbook函数
#这里是把str转换为Unicode类型
'''
print type(wb.active)   #<class 'openpyxl.worksheet.worksheet.Worksheet'>
print wb.active   #返回的是一个worksheet对象。<Worksheet "\u8282\u76ee\u8868">

print wb.read_only

print wb.encoding

print wb.worksheets

sheetnames = wb.get_sheet_names()     #返回一个包含所有表名的列表
print sheetnames
print type(sheetnames[0])
print sheetnames[0]

print wb.get_sheet_by_name(u'\u8282\u76ee\u8868')
'''

############表操作################

ws = wb.active    #ws是活跃表格的表对象
'''
print ws.title    #获得表格名
print ws.dimensions   #获得含有数据的表格大小。这里结果是A1:D211
print ws.max_row   #表的最大行。211
print ws.min_row
print ws.max_column   #表的最大列。4
print ws.min_column
'''

col = ws.values
#按行获取cell的内容。返回的是一个生成器对象。<generator object values at 0x00000000045873A8>
'''
for data in col:
    print data
    print data[2]
#data是列表。元素是每一列的内容。空的就为None。

row_gen = ws.rows   #按行获取cell对象。这也是一个生成器
for row in row_gen:
    print row

#(<Cell u'\u8282\u76ee\u8868'.A1>, <Cell u'\u8282\u76ee\u8868'.B1>, <Cell u'\u8282\u76ee\u8868'.C1>, <Cell u'\u8282\u76ee\u8868'.D1>)
#每一个row是一个元祖。元祖的元素是这一行的cell对象


###########单元格操作################

c3 = ws['C3']
#<Cell u'\u8282\u76ee\u8868'.C3>  c3是一个cell对象

c4 = ws.cell(row=4,column=3)
print c4
#也可以通过这种方式获得指定的cell对象<Cell u'\u8282\u76ee\u8868'.C4>

print c3.row   #3
print c3.column    #C
print c3.value    #07:50-10:15 WWE SmackDown第928期（英文解说）
print type(c3.value)   #<type 'unicode'>
#print c3.cordinate  好像没这个方法？
'''
#####打印出表格内容###########


#for row in ws.values:
#    print(*row)

'''
for i in range(ws.min_row,ws.max_row + 1):
    for j in range(ws.min_column,ws.max_column + 1):
        print ws.cell(row=i,column=j).value,
    print ''
'''

c4 = ws.cell(row=4,column=3)
print c4.value
