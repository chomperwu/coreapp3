#!/usr/bin/env python
# -*- coding: utf-8 -*-

#把指定excel文档中的内容筛选出来，生成一个新的文档

import openpyxl

old_wb = openpyxl.load_workbook(r'D:\pythonproject\coreapp3\excel编程\5月31日体育节目表.xlsx'.decode('utf-8'))

new_wb = openpyxl.Workbook()
#生成一个新的excel文档

old_ws = old_wb.get_sheet_by_name(u'节目表')
new_ws = new_wb.active

#for old_cell in old_ws.rows:
#    if old_cell[3].value == u'足球2（推流）':
#        pass

for i in range(old_ws.min_row,old_ws.max_row + 1):
    new_ws.cell(row=i + 1, column=1).value = old_ws.cell(row=i,column=4).value


new_ws.cell(row=1,column=1).value = '足球2（推流）'

new_ws['C3'] = 'Hello world!'
for i in range(10):
  new_ws["B%d" % (i+1)] = i + 1

new_wb.save('sample.xlsx')



